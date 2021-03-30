from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.utils.cell import absolute_coordinate
from caiman_functions.caiman_styles import caiman_excel_styles as ca_ex_st

# analyzed_cells function calculates the number of cells analyze during the experiment.


def analyzed_cell_number(sheet):
    new_file_max_column = sheet.max_column
    new_file_max_row = sheet.max_row
    analyzed_cells = 0
    for column in range(2, new_file_max_column + 1):
        if sheet.cell(row=new_file_max_row, column=column).value == None:
            pass
        else:
            analyzed_cells += 1
    return analyzed_cells


# average_se_trace_full_experiment function calculates the average and standard error of all the samples in each time-point
# generating and average trace for the whole experiment with its standard error.
# file_max_column: calculated by any of the analysis functions.
# file_max_row: calculated by any of the analysis functions.
# sheet: calculated by any of the analysis functions.


def average_se_trace_full_experiment(file_max_column, file_max_row, sheet):
    average_trace_header = sheet.cell(
        row=2, column=file_max_column + 4).coordinate
    sheet[average_trace_header] = "Experiment Average"
    ca_ex_st.style_headers(average_trace_header, sheet)

    average_trace_SE_header = sheet.cell(
        row=2, column=file_max_column + 5).coordinate
    sheet[average_trace_SE_header] = "Experiment SE"
    ca_ex_st.style_headers(average_trace_SE_header, sheet)

    for row in range(4, file_max_row + 1):
        first_cell = sheet.cell(row=row, column=3).coordinate
        last_cell = sheet.cell(row=row, column=file_max_column).coordinate
        average_cells = f"= average({first_cell}:{last_cell})"
        average_value = sheet.cell(
            row=row, column=file_max_column + 4).coordinate
        sheet[average_value] = average_cells
        ca_ex_st.style_number(average_value, sheet)
        standar_error_cells = f"= stdev({first_cell}:{last_cell})/sqrt({file_max_column-2})"
        standar_error_value = sheet.cell(
            row=row, column=file_max_column + 5).coordinate
        sheet[standar_error_value] = standar_error_cells
        ca_ex_st.style_number(standar_error_value, sheet)


# average_se_trace_full_experiment_chart generates a plot with the average values from the average_se_trace_full_experiment function.


def average_se_trace_full_experiment_chart(file_max_column, file_max_row, sheet):
    chart_cell = sheet.cell(row=4, column=file_max_column + 7).coordinate

    chart = ScatterChart()
    chart.style = 2
    chart.title = "Experiment average trace"
    chart.y_axis.title = "Fura2 fluorescence ratio (a.u)"
    chart.x_axis.title = "Time (s)"
    chart.legend = None
    chart.height = 10  # default is 7.5
    chart.width = 20  # default is 15
    chart.x_axis.majorUnit = 60
    ca_ex_st.style_chart(chart.title, chart)

    xvalues = Reference(sheet, min_col=file_max_column + 3, min_row=4,
                        max_col=file_max_column + 3, max_row=file_max_row)
    yvalues = Reference(sheet, min_col=file_max_column + 4, min_row=4,
                        max_col=file_max_column + 4, max_row=file_max_row)
    series = Series(yvalues, xvalues)
    series_trendline = Series(yvalues, xvalues)
    chart.series.append(series)
    chart.series.append(series_trendline)

    sheet.add_chart(chart, chart_cell)


# basal_ratio_pretreatment_and_increment_calculation function calculates the basal ratio before pre-estimuli or precalcium,
# and generates the "ratio true value - ratio basal value" for each frame,
# allowing us to calculate the agonist-evoked Ca2+ release and/or entry
# as the integral of the rise in fura-2 fluorescence ratio
# for 2½ min after the addition of the agonist or CaCl2.
# to achieve that, the function requieres the following values:
# column: calculated by any of the analysis functions.
# integral_time: Set as default as 150 seconds, but could be modified by the user.
# file_max_column: calculated by any of the analysis functions.
# row_number: use pre_stimuli to calculate the calcium release integral or pre_calcium for the calcium entry integral.
# sheet: calculated by any of the analysis functions.

def basal_ratio_pretreatment_and_increment_calculation(column, file_max_column, integral_time, row_number, sheet):
    ca_integral = row_number
    basal_begining = sheet.cell(
        row=ca_integral - 11, column=column).coordinate
    basal_ending = sheet.cell(
        row=ca_integral - 1, column=column).coordinate
    ca_integral_header = sheet.cell(
        row=ca_integral, column=file_max_column + 7).coordinate
    sheet[ca_integral_header] = f"= average({basal_begining}:{basal_ending})"
    ca_ex_st.style_average_values(ca_integral_header, sheet)

    # Cálculo incremento ratio de calcio basal tras el estímulo
    for row in range(ca_integral + 1, ca_integral + 1 + integral_time + 1):
        ratio_sample = sheet.cell(
            row=row, column=column).coordinate
        ratio_normalized = sheet.cell(
            row=row, column=file_max_column + 7).coordinate
        sheet[
            ratio_normalized] = f"= {ratio_sample}-{absolute_coordinate(ca_integral_header)}"
        ca_ex_st.style_number(ratio_normalized, sheet)


# experiment_time_in_seconds function generates a column with the experiment duration in seconds.
# adquisiton_time: Set as default as 2 seconds, but could be modified by the user.
# file_max_column: calculated by any of the analysis functions.
# file_max_row: calculated by any of the analysis functions.
# sheet: calculated by any of the analysis functions.


def calcium_oscillation_experiment_time_in_seconds(adquisiton_time, file_max_row, sheet):
    time_header = sheet.cell(row=1, column=1).coordinate
    sheet[time_header] = "Time (s)"
    ca_ex_st.style_headers(time_header, sheet)
    experiment_time = 0
    for row in range(2, file_max_row + 1):
        experiment_time_row = sheet.cell(
            row=row, column=1).coordinate
        sheet[experiment_time_row] = experiment_time
        ca_ex_st.style_time(experiment_time_row, sheet)
        experiment_time += adquisiton_time


# experiment_time_in_seconds function generates a column with the experiment duration in seconds.
# adquisiton_time: Set as default as 2 seconds, but could be modified by the user.
# file_max_column: calculated by any of the analysis functions.
# file_max_row: calculated by any of the analysis functions.
# sheet: calculated by any of the analysis functions.


def experiment_time_in_seconds(adquisiton_time, file_max_column, file_max_row, sheet):
    time_header = sheet.cell(row=2, column=file_max_column + 3).coordinate
    sheet[time_header] = "Experiment Time (s)"
    ca_ex_st.style_headers(time_header, sheet)
    experiment_time = 0
    for row in range(4, file_max_row + 1):
        experiment_time_row = sheet.cell(
            row=row, column=file_max_column + 3).coordinate
        sheet[experiment_time_row] = experiment_time
        ca_ex_st.style_time(experiment_time_row, sheet)
        experiment_time += adquisiton_time


# parameter_integral_ratio_values function generates an excel cell with the calculated integral value in the integral_time range
# for a given parameter in each sample. (Parameters= Calcium release | Calcium Entry)
# adquisiton_time: Set as default as 2 seconds, but could be modified by the user.
# file_max_column: calculated by any of the analysis functions.
# integral_time: Set as default as 150 seconds, but could be modified by the user.
# n_integral_value: value that calculates the position of the excel cell. Calculated by any of the analysis functions.
    # n_calcium_release_integral_value | n_calcium_entry_integral_value
# row_number: use pre_stimuli to calculate the calcium release integral or pre_calcium for the calcium entry integral.
# sheet: calculated by any of the analysis functions.


def parameter_integral_ratio_values(adquisiton_time, file_max_column, integral_time, n_integral_value, row_number, sheet):
    calcium_beginning = sheet.cell(
        row=row_number + 1, column=file_max_column + 7).coordinate
    calcium_ending = sheet.cell(
        row=row_number + 1 + integral_time, column=file_max_column + 7).coordinate
    calcium_integral = f"= (sum({calcium_beginning}:{calcium_ending})*{adquisiton_time})"
    cell_n_calcium_value = sheet.cell(
        row=n_integral_value, column=file_max_column + 7).coordinate
    sheet[cell_n_calcium_value] = calcium_integral
    ca_ex_st.style_number(cell_n_calcium_value, sheet)


# parameter_peak_values function generates an excel cell with the maximun value (Peak) in the integral_time range
# for a given parameter in each sample. (Parameters= Calcium release | Calcium Entry)
# file_max_column: calculated by any of the analysis functions.
# integral_time: Set as default as 150 seconds, but could be modified by the user.
# n_integral_value: value that calculates the position of the excel cell. Calculated by any of the analysis functions.
    # n_calcium_release_peak_value |  n_calcium_entry_peak_value
# row_number: use pre_stimuli to calculate the calcium release integral or pre_calcium for the calcium entry integral.
# sheet: calculated by any of the analysis functions.


def parameter_peak_values(file_max_column, integral_time, n_integral_value, row_number, sheet):
    calcium_beginning = sheet.cell(
        row=row_number + 1, column=file_max_column + 7).coordinate
    calcium_ending = sheet.cell(
        row=row_number + 1 + integral_time, column=file_max_column + 7).coordinate
    calcium_integral = f"= MAX({calcium_beginning}:{calcium_ending})"
    cell_n_calcium_value = sheet.cell(
        row=n_integral_value, column=file_max_column + 7).coordinate
    sheet[cell_n_calcium_value] = calcium_integral
    ca_ex_st.style_number(cell_n_calcium_value, sheet)


# # slope_values function generates an excel cell with the slope value of the trendline calculated from the integral_time range
# for a given parameter in each sample. (Parameters= Calcium release | Calcium Entry)
# # file_max_column: calculated by any of the analysis functions.
# # integral_time: Set as default as 150 seconds, but could be modified by the user.
# # n_integral_value: value that calculates the position of the excel cell. Calculated by any of the analysis functions.
# #     n_calcium_release_slope_value | n_calcium_entry_slope_value
# # row_number: use pre_stimuli to calculate the calcium release integral or pre_calcium for the calcium entry integral.
# # sheet: calculated by any of the analysis functions.

def parameter_slope_values(column, file_max_column, n_integral_value, row_number, sheet, slope_time, time_column):
    calcium_beginning = sheet.cell(
        row=row_number + 1, column=column).coordinate
    calcium_ending = sheet.cell(
        row=row_number + 1 + slope_time, column=column).coordinate
    slope_beginning = sheet.cell(
        row=row_number + 1, column=time_column).coordinate
    slope_ending = sheet.cell(
        row=row_number + 1 + slope_time, column=time_column).coordinate
    calcium_slope = f"= SLOPE({calcium_beginning}:{calcium_ending},{slope_beginning}:{slope_ending})"
    cell_n_calcium_value = sheet.cell(
        row=n_integral_value, column=file_max_column + 7).coordinate
    sheet[cell_n_calcium_value] = calcium_slope
    ca_ex_st.style_number(cell_n_calcium_value, sheet)


# parameters_header_position function generates an excel cell with the name (header) of the calculated parameter in each sample.
# file_max_column: calculated by any of the analysis functions.
# header: the string that will appear in the excel cell: "N Release","Peak Release", "Slope Release", "N Entry", "Peak Entry", "Slope Entry".
# n_integral_title: value that calculates the position of the excel cell. Calculated by any of the analysis functions.
    # n_calcium_release_integral_title, n_calcium_release_peak_title, n_calcium_release_slope_title
    # n_calcium_entry_integral_title, n_calcium_entry_peak_title, n_calcium_entry_slope_title
# sheet: calculated by any of the analysis functions.


def parameters_header_position(file_max_column, header, n_integral_title, sheet):
    cell_n_calcium_title = sheet.cell(
        row=n_integral_title, column=file_max_column + 7).coordinate
    sheet[cell_n_calcium_title] = f"{header}"
    ca_ex_st.style_average_values(cell_n_calcium_title, sheet)


# single_cell_average_se_paramemeters function calculates the average and standard error of each parameter for the full experiment.
# analyzed_parameter: Release | Release Peak | Release Slope | Entry | Entry Peak | Entry Slope.
# n_integral_title: value that calculates the position of the excel cell. Calculated by any of the analysis functions.
#     n_calcium_release_integral_title | n_calcium_entry_integral_title
#     n_calcium_release_peak_title | n_calcium_entry_peak_title
#     n_calcium_release_slope_title | n_calcium_entry_slope_title
# n_integral_value: value that calculates the position of the excel cell. Calculated by any of the analysis functions.
#     n_calcium_release_integral_value | n_calcium_entry_integral_value
#     n_calcium_release_peak_value | n_calcium_entry_peak_value
#     n_calcium_release_slope_value | n_calcium_entry_slope_value
# sheet: calculated by any of the analysis functions.
# total_column: final number of columns in the file after all the calculations are performed.


def single_cell_average_se_paramemeters(analyzed_cells, analyzed_parameter, n_integral_title, n_integral_value,  sheet, total_column):
    average_title = sheet.cell(
        row=n_integral_title, column=total_column + 2).coordinate
    sheet[average_title] = f"Average  {analyzed_parameter}"
    ca_ex_st.style_average_values(average_title, sheet)
    standar_error_title = sheet.cell(
        row=n_integral_title, column=total_column + 3).coordinate
    sheet[standar_error_title] = "SE"
    ca_ex_st.style_average_values(standar_error_title, sheet)

    first_cell = sheet.cell(
        row=n_integral_value, column=(total_column + 1) - analyzed_cells).coordinate
    last_cell = sheet.cell(
        row=n_integral_value, column=total_column).coordinate
    average_cells = f"= average({first_cell}:{last_cell})"
    average_value = sheet.cell(
        row=n_integral_value, column=total_column + 2).coordinate
    sheet[average_value] = average_cells
    ca_ex_st.style_number(average_value, sheet)

    standar_error_cells = f"= stdev({first_cell}:{last_cell})/sqrt({analyzed_cells})"
    standar_error_value = sheet.cell(
        row=n_integral_value, column=total_column + 3).coordinate
    sheet[standar_error_value] = standar_error_cells
    ca_ex_st.style_number(standar_error_value, sheet)


# single_cell_ratio_normalized_f_f0 function normalizes the ratio along the time of each sample with its basal value.
# To achieve that, the program, and not this function, will create a copy of the main worksheet 'Time Measurement Report' and rename it as 'F_F0'.
# Following along the function will them calculate the average ratio basal value for the 10 first values
# within the timelap in 'Time Measurement Report', and will write the result in the first row of 'F_F0' for each sample.
# Finally, it will divide each of the values in the timelap for each sample with its calculated average,
# writing the result in the worksheet 'F_F0'.


def single_cell_ratio_normalized_f_f0(sheet_f_f0_max_column, sheet_f_f0_max_row, sheet_f_f0):
    number_cell = 1
    for column in range(1, sheet_f_f0_max_column + 1):
        if "Ratio" in sheet_f_f0.cell(row=2, column=column).value:
            sample_number_header = sheet_f_f0.cell(
                row=2, column=column).coordinate
            sheet_f_f0[sample_number_header] = f"F/F0 - Cell {number_cell}"
            ca_ex_st.style_headers(sample_number_header, sheet_f_f0)

            basal_begining = sheet_f_f0.cell(
                row=4, column=column).coordinate
            basal_ending = sheet_f_f0.cell(
                row=14, column=column).coordinate
            ca_f_f0_integral_header = sheet_f_f0.cell(
                row=1, column=column).coordinate
            sheet_f_f0[
                ca_f_f0_integral_header] = f"= average('Time Measurement Report'!{basal_begining}:{basal_ending})"
            ca_ex_st.style_average_values(ca_f_f0_integral_header, sheet_f_f0)

            # Cálculo ratio de calcio basal al comienzo del experimento
            for row in range(4, sheet_f_f0_max_row + 1):
                ratio_sample = sheet_f_f0.cell(
                    row=row, column=column).coordinate
                ratio_normalized = sheet_f_f0.cell(
                    row=row, column=column).coordinate
                sheet_f_f0[
                    ratio_normalized] = f"= 'Time Measurement Report'!{ratio_sample}/F_F0!{absolute_coordinate(ca_f_f0_integral_header)}"
                ca_ex_st.style_number(ratio_normalized, sheet_f_f0)

            number_cell += 1


# single_cell_traces_in_one_chart function generates 1 single scatter chart within the file with all the traces represented where:
#   x_axis = time(s)
#   y_axis = Fura2 fluorescence.
#   series = one serie for each analyzed cell
# file_max_column: calculated by any of the analysis functions.
# file_max_row: calculated by any of the analysis functions.
# sheet: calculated by any of the analysis functions.

def single_cell_traces_in_one_chart(file_max_column, file_max_row, sheet):
    chart_cell = sheet.cell(row=25, column=file_max_column + 7).coordinate

    chart = ScatterChart()
    chart.style = 2
    chart.title = "Single cell traces"
    chart.y_axis.title = "Fura2 fluorescence ratio (a.u)"
    chart.x_axis.title = "Time (s)"
    chart.legend = None
    chart.height = 10  # default is 7.5
    chart.width = 20  # default is 15
    chart.x_axis.majorUnit = 60
    ca_ex_st.style_chart(chart.title, chart)

    xvalues = Reference(sheet, min_col=file_max_column + 3, min_row=4,
                        max_col=file_max_column + 3, max_row=file_max_row)

    for column in range(3, file_max_column + 1):
        # print(column)
        values = Reference(sheet, min_col=column,
                           min_row=4, max_row=file_max_row)
        series = Series(values, xvalues)
        chart.series.append(series)

    sheet.add_chart(chart, chart_cell)

# single_cell_trace_in_individual_chart function generates 1 scatter chart within the file for each of the traces where:
#   x_axis = time(s)
#   y_axis = Fura2 fluorescence.#
# column_individual_trace_charts: Determines the column where the chart will be created
# experiment_number: Used as the chart title.
# file_max_row: calculated by any of the analysis functions.
# row_individual_trace_charts: Determines the column where the chart will be created
# sheet: calculated by any of the analysis functions.
# time_column: calculates the maximun column number within the file.


def single_cell_trace_in_individual_chart(column, column_individual_trace_charts, experiment_number, file_max_row, sheet, row_individual_trace_charts, time_column):

    chart_cell = sheet.cell(row=row_individual_trace_charts,
                            column=column_individual_trace_charts).coordinate

    chart = ScatterChart()
    chart.style = 2
    chart.title = f"Cell {experiment_number}: individual_trace"
    chart.y_axis.title = "Fura2 fluorescence ratio (a.u)"
    chart.x_axis.title = "Time (s)"
    chart.legend = None
    chart.height = 7.5  # default is 7.5
    chart.width = 15  # default is 15
    chart.x_axis.majorUnit = 60
    ca_ex_st.style_chart(chart.title, chart)

    xvalues = Reference(sheet, min_col=time_column, min_row=4,
                        max_col=time_column, max_row=file_max_row)
    yvalues = Reference(sheet, min_col=column, min_row=4,
                        max_col=column, max_row=file_max_row)
    series = Series(yvalues, xvalues)
    chart.series.append(series)

    sheet.add_chart(chart, chart_cell)


# single_cell_slope_trace_chart function generates 1 scatter chart within the file for each of the traces where:
#   x_axis = slope_time
#   y_axis = Fura2 fluorescence.#
# column_individual_trace_charts: Determines the column where the chart will be created
# experiment_number: Used as the chart title.
# file_max_row: calculated by any of the analysis functions.
# row_individual_trace_charts: Determines the column where the chart will be created


def single_cell_slope_trace_chart(column, column_slope_charts, experiment_number, row_charts, row_number, sheet, slope_name, slope_time, time_column):

    chart_cell = sheet.cell(row=row_charts,
                            column=column_slope_charts).coordinate

    chart = ScatterChart()
    chart.style = 2
    chart.title = f"Cell {experiment_number}: {slope_name} slope"
    chart.y_axis.title = "Fura2 fluorescence ratio (a.u)"
    chart.x_axis.title = "Time (s)"
    chart.legend = None
    chart.height = 7.5  # default is 7.5
    chart.width = 15  # default is 15
    chart.x_axis.majorUnit = 10
    ca_ex_st.style_chart(chart.title, chart)

    xvalues = Reference(sheet, min_col=time_column, min_row=row_number + 1,
                        max_col=time_column, max_row=row_number + 1 + slope_time)
    yvalues = Reference(sheet, min_col=column, min_row=row_number + 1,
                        max_col=column, max_row=row_number + 1 + slope_time)
    series = Series(yvalues, xvalues)
    series_trendline = Series(yvalues, xvalues)
    chart.series.append(series)
    chart.series.append(series_trendline)

    line = chart.series[0]
    line.graphicalProperties.line.noFill = True
    line.trendline = Trendline(dispRSqr=True, dispEq=True)

    sheet.add_chart(chart, chart_cell)
