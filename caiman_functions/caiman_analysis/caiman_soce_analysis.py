# Hay que meter las fórmulas en inglés

import os
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.title import Title
from openpyxl.chart.text import RichText
from openpyxl.chart.trendline import Trendline
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties, RegularTextRun
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.cell import absolute_coordinate


from caiman_functions.caiman_analysis import caiman_calculation as ca_cal
from caiman_functions.caiman_styles import caiman_excel_styles as ca_ex_st


def soce_analysis(adquisiton_time, integral_time, keyword, pre_calcium, pre_stimuli, route, slope_time_entry, slope_time_release):
    wb = openpyxl.load_workbook(route)
    sheet = wb.active
    if "imageJ" not in route.stem:
        ca_cal.initial_sheet_formating(route.stem, sheet, wb)

    file_max_column = sheet.max_column
    file_max_row = sheet.max_row
    time_column = file_max_column + 3
    sheet_f_f0_max_column = sheet.max_column
    sheet_f_f0_max_row = sheet.max_row
    n_calcium_release_integral_title = file_max_row + 3
    n_calcium_release_integral_value = file_max_row + 4
    n_calcium_release_peak_title = file_max_row + 5
    n_calcium_release_peak_value = file_max_row + 6
    n_calcium_release_slope_title = file_max_row + 7
    n_calcium_release_slope_value = file_max_row + 8

    n_calcium_entry_integral_title = file_max_row + 10
    n_calcium_entry_integral_value = file_max_row + 11
    n_calcium_entry_peak_title = file_max_row + 12
    n_calcium_entry_peak_value = file_max_row + 13
    n_calcium_entry_slope_title = file_max_row + 14
    n_calcium_entry_slope_value = file_max_row + 15

    ca_cal.experiment_time_in_seconds(
        adquisiton_time, file_max_column, file_max_row, sheet)

    ca_cal.average_se_trace_full_experiment(
        file_max_column, file_max_row, sheet)

    ca_cal.average_se_trace_full_experiment_chart(
        file_max_column, file_max_row, sheet)

    ca_cal.single_cell_traces_in_one_chart(
        file_max_column, file_max_row, sheet)

    experiment_title_cell_header = sheet.cell(row=1, column=1).coordinate
    first_cell_header = sheet.cell(row=2, column=1).coordinate
    ca_ex_st.style_headers(first_cell_header, sheet)
    ca_ex_st.style_experiment_title_header(experiment_title_cell_header, sheet)

    number_cell = 1
    for column in range(1, file_max_column + 1):
        if sheet.cell(row=2, column=column).value:
            if keyword in sheet.cell(row=2, column=column).value:
                sample_cell_header = sheet.cell(
                    row=2, column=column).coordinate
                sheet[sample_cell_header] = f"#{number_cell:02} {keyword}"
                ca_ex_st.style_headers(sample_cell_header, sheet)
                sample_number_header = sheet.cell(
                    row=2, column=file_max_column + 7).coordinate
                sheet[sample_number_header] = f"Cell {number_cell:02}"
                ca_ex_st.style_headers(sample_number_header, sheet)

                ca_cal.basal_ratio_pretreatment_and_increment_calculation(column, file_max_column,
                                                                          integral_time, pre_stimuli, sheet)
                ca_cal.basal_ratio_pretreatment_and_increment_calculation(column, file_max_column,
                                                                          integral_time, pre_calcium, sheet)

                ca_cal.parameters_header_position(file_max_column, "Release Integral",
                                                  n_calcium_release_integral_title, sheet)

                ca_cal.parameter_integral_ratio_values(adquisiton_time, file_max_column, integral_time,
                                                       n_calcium_release_integral_value, pre_stimuli, sheet)

                ca_cal.parameters_header_position(file_max_column, "Release Peak",
                                                  n_calcium_release_peak_title, sheet)

                ca_cal.parameter_peak_values(file_max_column, integral_time,
                                             n_calcium_release_peak_value, pre_stimuli, sheet)

                ca_cal.parameters_header_position(file_max_column, "Release Slope",
                                                  n_calcium_release_slope_title, sheet)

                ca_cal.parameter_slope_values(column, file_max_column,
                                              n_calcium_release_slope_value, pre_stimuli, sheet, slope_time_release, time_column)

                ca_cal.parameters_header_position(file_max_column, "Entry Integral",
                                                  n_calcium_entry_integral_title, sheet)

                ca_cal.parameter_integral_ratio_values(adquisiton_time, file_max_column, integral_time,
                                                       n_calcium_entry_integral_value, pre_calcium, sheet)

                ca_cal.parameters_header_position(file_max_column, "Entry Peak",
                                                  n_calcium_entry_peak_title, sheet)

                ca_cal.parameter_peak_values(file_max_column, integral_time,
                                             n_calcium_entry_peak_value, pre_calcium, sheet)

                ca_cal.parameters_header_position(file_max_column, "Entry Slope",
                                                  n_calcium_entry_slope_title, sheet)

                ca_cal.parameter_slope_values(column, file_max_column,
                                              n_calcium_entry_slope_value, pre_calcium, sheet, slope_time_entry, time_column)

                number_cell += 1
                file_max_column += 1

    print("")
    print(f"Number of analyzed cells: {ca_cal.analyzed_cell_number(sheet)}.")
    analyzed_cells = ca_cal.analyzed_cell_number(sheet)
    total_column = sheet.max_column

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Release", n_calcium_release_integral_title,
                                               n_calcium_release_integral_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Release Peak", n_calcium_release_peak_title,
                                               n_calcium_release_peak_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Release Slope", n_calcium_release_slope_title,
                                               n_calcium_release_slope_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Entry", n_calcium_entry_integral_title,
                                               n_calcium_entry_integral_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Entry Peak", n_calcium_entry_peak_title,
                                               n_calcium_entry_peak_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Entry Slope", n_calcium_entry_slope_title,
                                               n_calcium_entry_slope_value,  sheet, total_column)

    column_individual_trace_charts = sheet.max_column + 1
    release_slope_charts = sheet.max_column + 9
    entry_slope_charts = sheet.max_column + 17
    experiment_number = 1
    row_charts = 4

    for column in range(1, file_max_column + 1):
        if sheet.cell(row=2, column=column).value:
            if keyword in sheet.cell(row=2, column=column).value:
                ca_cal.single_cell_trace_in_individual_chart(
                    column, column_individual_trace_charts, sheet.cell(row=2, column=column).value, file_max_row, sheet, row_charts, time_column)
                ca_cal.single_cell_slope_trace_chart(column, release_slope_charts, sheet.cell(row=2, column=column).value,
                                                     row_charts, pre_stimuli, sheet, "Release", slope_time_release, time_column)
                ca_cal.single_cell_slope_trace_chart(column, entry_slope_charts, sheet.cell(row=2, column=column).value,
                                                     row_charts, pre_calcium, sheet,  "Entry", slope_time_entry, time_column)
                row_charts += 15

    ca_ex_st.colum_max_widths_non_oscillation_calcium_experiments(sheet)
    wb.copy_worksheet(sheet)
    sheet_f_f0 = wb["Time Measurement Report Copy"]
    sheet_f_f0.title = "F_F0"
    wb.active = sheet_f_f0
    ca_cal.single_cell_ratio_normalized_f_f0(
        keyword, sheet_f_f0_max_column, sheet_f_f0_max_row, sheet_f_f0)
    ca_cal.average_se_trace_full_experiment_chart(
        sheet_f_f0_max_column, sheet_f_f0_max_row, sheet_f_f0)
    ca_cal.single_cell_traces_in_one_chart(sheet_f_f0_max_column,
                                           sheet_f_f0_max_row, sheet_f_f0)

    row_charts_f_f0 = 4
    for column in range(1, sheet_f_f0_max_column + 1):
        if sheet_f_f0.cell(row=2, column=column).value:
            if "F/F0" in sheet_f_f0.cell(row=2, column=column).value:
                ca_cal.single_cell_trace_in_individual_chart(
                    column, column_individual_trace_charts, sheet_f_f0.cell(row=2, column=column).value, file_max_row, sheet_f_f0, row_charts_f_f0, time_column)
                ca_cal.single_cell_slope_trace_chart(column, release_slope_charts, sheet_f_f0.cell(row=2, column=column).value,
                                                     row_charts_f_f0, pre_stimuli, sheet_f_f0, "Release", slope_time_release, time_column)
                ca_cal.single_cell_slope_trace_chart(column, entry_slope_charts, sheet_f_f0.cell(row=2, column=column).value,
                                                     row_charts_f_f0, pre_calcium, sheet_f_f0, "Entry", slope_time_entry, time_column)
                row_charts_f_f0 += 15

    wb.active = sheet
    wb.save(f"{route.stem}_analyzed.xlsx")
    print(f"{route.stem}_analyzed.xlsx has been saved.")


def soce_multi_analysis(adquisiton_time, file, folder, integral_time, keyword, pre_calcium, pre_stimuli, slope_time_entry, slope_time_release):
    route = os.path.join(folder, file)
    wb = openpyxl.load_workbook(route)
    sheet = wb.active
    if "imageJ" not in file:
        ca_cal.initial_sheet_formating(file, sheet, wb)

    file_max_column = sheet.max_column
    file_max_row = sheet.max_row
    time_column = file_max_column + 3
    sheet_f_f0_max_column = sheet.max_column
    sheet_f_f0_max_row = sheet.max_row
    n_calcium_release_integral_title = file_max_row + 3
    n_calcium_release_integral_value = file_max_row + 4
    n_calcium_release_peak_title = file_max_row + 5
    n_calcium_release_peak_value = file_max_row + 6
    n_calcium_release_slope_title = file_max_row + 7
    n_calcium_release_slope_value = file_max_row + 8

    n_calcium_entry_integral_title = file_max_row + 10
    n_calcium_entry_integral_value = file_max_row + 11
    n_calcium_entry_peak_title = file_max_row + 12
    n_calcium_entry_peak_value = file_max_row + 13
    n_calcium_entry_slope_title = file_max_row + 14
    n_calcium_entry_slope_value = file_max_row + 15

    ca_cal.experiment_time_in_seconds(
        adquisiton_time, file_max_column, file_max_row, sheet)

    ca_cal.average_se_trace_full_experiment(
        file_max_column, file_max_row, sheet)

    ca_cal.average_se_trace_full_experiment_chart(
        file_max_column, file_max_row, sheet)

    ca_cal.single_cell_traces_in_one_chart(
        file_max_column, file_max_row, sheet)

    experiment_title_cell_header = sheet.cell(row=1, column=1).coordinate
    first_cell_header = sheet.cell(row=2, column=1).coordinate
    ca_ex_st.style_headers(first_cell_header, sheet)
    ca_ex_st.style_experiment_title_header(experiment_title_cell_header, sheet)

    number_cell = 1
    for column in range(1, file_max_column + 1):
        if sheet.cell(row=2, column=column).value:
            if keyword in sheet.cell(row=2, column=column).value:
                sample_cell_header = sheet.cell(
                    row=2, column=column).coordinate
                sheet[sample_cell_header] = f"#{number_cell:02} {keyword}"
                ca_ex_st.style_headers(sample_cell_header, sheet)
                sample_number_header = sheet.cell(
                    row=2, column=file_max_column + 7).coordinate
                sheet[sample_number_header] = f"Cell {number_cell:02}"
                ca_ex_st.style_headers(sample_number_header, sheet)

                ca_cal.basal_ratio_pretreatment_and_increment_calculation(column, file_max_column,
                                                                          integral_time, pre_stimuli, sheet)
                ca_cal.basal_ratio_pretreatment_and_increment_calculation(column, file_max_column,
                                                                          integral_time, pre_calcium, sheet)

                ca_cal.parameters_header_position(file_max_column, "Release Integral",
                                                  n_calcium_release_integral_title, sheet)

                ca_cal.parameter_integral_ratio_values(adquisiton_time, file_max_column, integral_time,
                                                       n_calcium_release_integral_value, pre_stimuli, sheet)

                ca_cal.parameters_header_position(file_max_column, "Release Peak",
                                                  n_calcium_release_peak_title, sheet)

                ca_cal.parameter_peak_values(file_max_column, integral_time,
                                             n_calcium_release_peak_value, pre_stimuli, sheet)

                ca_cal.parameters_header_position(file_max_column, "Release Slope",
                                                  n_calcium_release_slope_title, sheet)

                ca_cal.parameter_slope_values(column, file_max_column, n_calcium_release_slope_value,
                                              pre_stimuli, sheet, slope_time_release, time_column)

                ca_cal.parameters_header_position(file_max_column, "Entry Integral",
                                                  n_calcium_entry_integral_title, sheet)

                ca_cal.parameter_integral_ratio_values(adquisiton_time, file_max_column, integral_time,
                                                       n_calcium_entry_integral_value, pre_calcium, sheet)

                ca_cal.parameters_header_position(file_max_column, "Entry Peak",
                                                  n_calcium_entry_peak_title, sheet)

                ca_cal.parameter_peak_values(file_max_column, integral_time,
                                             n_calcium_entry_peak_value, pre_calcium, sheet)

                ca_cal.parameters_header_position(file_max_column, "Entry Slope",
                                                  n_calcium_entry_slope_title, sheet)

                ca_cal.parameter_slope_values(column, file_max_column,
                                              n_calcium_entry_slope_value, pre_calcium, sheet, slope_time_entry, time_column)

                number_cell += 1
                file_max_column += 1

    print("")
    print(f"Number of analyzed cells: {ca_cal.analyzed_cell_number(sheet)}.")
    analyzed_cells = ca_cal.analyzed_cell_number(sheet)
    total_column = sheet.max_column

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Release", n_calcium_release_integral_title,
                                               n_calcium_release_integral_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Release Peak", n_calcium_release_peak_title,
                                               n_calcium_release_peak_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Release Slope", n_calcium_release_slope_title,
                                               n_calcium_release_slope_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Entry", n_calcium_entry_integral_title,
                                               n_calcium_entry_integral_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Entry Peak", n_calcium_entry_peak_title,
                                               n_calcium_entry_peak_value,  sheet, total_column)

    ca_cal.single_cell_average_se_paramemeters(analyzed_cells, "Entry Slope", n_calcium_entry_slope_title,
                                               n_calcium_entry_slope_value,  sheet, total_column)

    column_individual_trace_charts = sheet.max_column + 1
    release_slope_charts = sheet.max_column + 9
    entry_slope_charts = sheet.max_column + 17
    row_charts = 4
    for column in range(1, file_max_column + 1):
        if sheet.cell(row=2, column=column).value:
            if keyword in sheet.cell(row=2, column=column).value:
                ca_cal.single_cell_trace_in_individual_chart(
                    column, column_individual_trace_charts, sheet.cell(row=2, column=column).value, file_max_row, sheet, row_charts, time_column)
                ca_cal.single_cell_slope_trace_chart(column, release_slope_charts, sheet.cell(row=2, column=column).value,
                                                     row_charts, pre_stimuli, sheet, "Release", slope_time_release, time_column)
                ca_cal.single_cell_slope_trace_chart(column, entry_slope_charts, sheet.cell(row=2, column=column).value,
                                                     row_charts, pre_calcium, sheet,  "Entry", slope_time_entry, time_column)
                row_charts += 15

    ca_ex_st.colum_max_widths_non_oscillation_calcium_experiments(sheet)
    wb.copy_worksheet(sheet)
    sheet_f_f0 = wb["Time Measurement Report Copy"]
    sheet_f_f0.title = "F_F0"
    wb.active = sheet_f_f0
    ca_cal.single_cell_ratio_normalized_f_f0(
        keyword, sheet_f_f0_max_column, sheet_f_f0_max_row, sheet_f_f0)
    ca_cal.average_se_trace_full_experiment_chart(
        sheet_f_f0_max_column, sheet_f_f0_max_row, sheet_f_f0)
    ca_cal.single_cell_traces_in_one_chart(sheet_f_f0_max_column,
                                           sheet_f_f0_max_row, sheet_f_f0)

    row_charts_f_f0 = 4
    for column in range(1, sheet_f_f0_max_column + 1):
        if sheet_f_f0.cell(row=2, column=column).value:
            if "F/F0" in sheet_f_f0.cell(row=2, column=column).value:
                ca_cal.single_cell_trace_in_individual_chart(
                    column, column_individual_trace_charts, sheet_f_f0.cell(row=2, column=column).value, file_max_row, sheet_f_f0, row_charts_f_f0, time_column)
                ca_cal.single_cell_slope_trace_chart(column, release_slope_charts, sheet_f_f0.cell(row=2, column=column).value,
                                                     row_charts_f_f0, pre_stimuli, sheet_f_f0, "Release", slope_time_release, time_column)
                ca_cal.single_cell_slope_trace_chart(column, entry_slope_charts, sheet_f_f0.cell(row=2, column=column).value,
                                                     row_charts_f_f0, pre_calcium, sheet_f_f0, "Entry", slope_time_entry, time_column)
                row_charts_f_f0 += 15

    wb.active = sheet
    splitted_file = file.split(".xlsx")[0]
    save_route = os.path.join(folder, splitted_file)
    wb.save(f"{save_route}_analyzed.xlsx")
    print(f"{splitted_file}_analyzed.xlsx has been saved.")
