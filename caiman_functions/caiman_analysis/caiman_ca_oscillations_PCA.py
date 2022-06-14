import matplotlib.pyplot as plt
import matplotlib.style
import numpy as np
import openpyxl
import os
import pandas as pd
import shutil

from matplotlib.ticker import MultipleLocator, FormatStrFormatter, AutoMinorLocator
from numpy import trapz
from pathlib import Path
from scipy import stats
from scipy.signal import find_peaks, peak_widths
from scipy.integrate import simps

from caiman_functions.caiman_analysis import caiman_calculation as ca_cal
from caiman_functions.caiman_styles import caiman_excel_styles as ca_ex_st
from caiman_functions.caiman_utilities.df_to_excel import append_df_to_excel, addapted_df_to_excel


def imaging_ca_analysis_PCA(adquisiton_time, keyword, peak_amplitude, peak_longitude, route, time_initial_linregress, y_max_value, y_min_value):
    converted_route = Path(route).parent
    os.chdir(converted_route)
    wb = openpyxl.load_workbook(route)

    sheet = wb.active
    # sheet.delete_cols(2)
    # sheet.delete_rows(3)
    sheet.delete_rows(1)

    file_max_row = sheet.max_row
    file_max_column = sheet.max_column
    ca_cal.calcium_oscillation_experiment_time_in_seconds(
        adquisiton_time, file_max_row, sheet)

    number_cell = 1
    for column in range(1, file_max_column + 1):
        if sheet.cell(row=1, column=column).value:
            if keyword in sheet.cell(row=1, column=column).value:
                sample_number_header = sheet.cell(
                    row=1, column=column).coordinate
                sheet[sample_number_header] = f"#{number_cell:02} {keyword}"
                ca_ex_st.style_headers(sample_number_header, sheet)
                number_cell += 1

    splitted_file = Path(route).stem
    non_space_filename = splitted_file.replace(" ", "_")
    wb.save(f"{non_space_filename}_modified.xlsx")

    file_route = f"{non_space_filename}_modified.xlsx"
    df = pd.read_excel(file_route)
    x = df["Time (s)"]
    y = df.iloc[:, 1:]
    max_column_num = len(df.columns)
    column_num_plots = max_column_num + 2
    row_num_plots = 2
    column_num_data = max_column_num + 10
    row_num_data = 2

    for column in y.columns:
        fig, ax = plt.subplots(constrained_layout=True)
        fig = matplotlib.pyplot.gcf()
        fig.set_size_inches(6, 3)
        ax.plot(x, y[column], color="black")

        #  Añade peaks al archivo excel
        peaks, _ = find_peaks(
            y[column], prominence=peak_amplitude, width=peak_longitude)      # BEST!
        parameter_value_peaks = f"Peak time {column}"
        df_peaks = pd.DataFrame(
            {parameter_value_peaks: peaks * adquisiton_time})
        append_df_to_excel(file_route, df_peaks.T, sheet_name="Time Measurement Report",
                           startrow=row_num_data, startcol=column_num_data, header=True, index=True, truncate_sheet=False)
        row_num_data += 2

        peaks, properties = find_peaks(
            y[column], height=0, prominence=peak_amplitude, width=peak_longitude)
        # properties["prominences"], properties["widths"]
        parameter_value_max_ratio = f"Max ratio {column}"
        df_max = pd.DataFrame(
            {parameter_value_max_ratio: properties["peak_heights"]})
        append_df_to_excel(file_route, df_max.T, sheet_name="Time Measurement Report",
                           startrow=row_num_data, startcol=column_num_data, header=False, index=True, truncate_sheet=False)
        row_num_data += 1

        parameter_value_max_prominence = f"Max prominence {column}"
        df_max = pd.DataFrame(
            {parameter_value_max_prominence: properties["prominences"]})
        append_df_to_excel(file_route, df_max.T, sheet_name="Time Measurement Report",
                           startrow=row_num_data, startcol=column_num_data, header=False, index=True, truncate_sheet=False)
        row_num_data += 1

        parameter_value_width = f"Width (sec) {column}"
        df_width = pd.DataFrame(
            {parameter_value_width: properties["widths"] * adquisiton_time})
        append_df_to_excel(file_route, df_width.T, sheet_name="Time Measurement Report",
                           startrow=row_num_data, startcol=column_num_data, header=False, index=True, truncate_sheet=False)

        row_num_data += 1
        right_peak_widths = properties["right_ips"] - peaks
        half_calcium_decrease_time = f"Time for half calcium decrease (sec) {column}"
        df_half_calcium_decrease_time = pd.DataFrame(
            {half_calcium_decrease_time: right_peak_widths * adquisiton_time})
        append_df_to_excel(file_route, df_half_calcium_decrease_time.T, sheet_name="Time Measurement Report",
                           startrow=row_num_data, startcol=column_num_data, header=False, index=True, truncate_sheet=False)

        row_num_data += 2

        # The numpy and scipy libraries include the composite trapezoidal (numpy.trapz)
        # and Simpson's (scipy.integrate.simps)rules.
        #  https://stackoverflow.com/questions/13320262/calculating-the-area-under-a-curve-given-a-set-of-coordinates-without-knowing-t/13323861#13323861
        #  Añade los valores de la recta de regresión obtenida con los valores iniciales y el área bajo la curva al archivo excel

        df_initial_time = x[:time_initial_linregress]
        df_initial_trend = y[column][:time_initial_linregress]
        res_initial = stats.linregress(
            df_initial_time, df_initial_trend)
        df_initial_linear_regresion = res_initial.intercept + res_initial.slope*x
        corrected_y_initial = y[column].sub(
            df_initial_linear_regresion.squeeze())
        r_squared_initial = float(f"{res_initial.rvalue**2:.6f}")
        intercept_initial = res_initial.intercept
        slope_initial = res_initial.slope
        addapted_df_to_excel(column, r_squared_initial, file_route,
                             column_num_data, row_num_data, "R-squared initial")
        row_num_data += 1
        addapted_df_to_excel(column, intercept_initial, file_route,
                             column_num_data, row_num_data, "Intercept initial")
        row_num_data += 1
        addapted_df_to_excel(column, slope_initial, file_route,
                             column_num_data, row_num_data, "Slope initial")
        row_num_data += 1
        area_initial_trapz = trapz(
            corrected_y_initial, x) * adquisiton_time
        addapted_df_to_excel(column, area_initial_trapz, file_route, column_num_data,
                             row_num_data, "Trapz AUC Initial linear regresion")
        row_num_data += 1
        area_initial_simps = simps(
            corrected_y_initial, x) * adquisiton_time
        addapted_df_to_excel(column, area_initial_simps, file_route, column_num_data,
                             row_num_data, "Simps AUC Initial linear regresion")
        row_num_data += 6

        # Make a plot with major ticks that are multiples of 20 and minor ticks that
        # are multiples of 5.  Label major ticks with '%d' formatting but don't label
        # minor ticks.

        ax.fill_between(
            x, y[column], df_initial_linear_regresion, color="peachpuff", where=x > time_initial_linregress)
        ax.xaxis.set_major_locator(MultipleLocator(60))
        ax.xaxis.set_minor_locator(MultipleLocator(30))
        ax.set_title(f'{column}', fontsize=12)
        ax.set_xlabel('Time (s)', fontsize=9)
        ax.set_ylabel('Fura2 fluorescencia (a.u)', fontsize=9)
        if y_min_value != "" and y_max_value != "":
            plt.ylim(y_min_value, y_max_value)
        else:
            pass

        ax.vlines(x=peaks*adquisiton_time, ymin=y[column][peaks] - properties["prominences"],
                  ymax=y[column][peaks], color="blue")
        ax.hlines(y=properties["width_heights"], xmin=properties["left_ips"]*adquisiton_time,
                  xmax=properties["right_ips"]*adquisiton_time, color="blue")

        ax.plot(peaks*adquisiton_time, y[column]
                [peaks], "o", color="red")

        ax.plot(x, df_initial_linear_regresion,
                color="saddlebrown", label="AUC Initial linear regresion")

        ax.legend()
        plt.rcParams.update({'figure.max_open_warning': 0})

        temporal_images = os.path.join(converted_route, non_space_filename)
        Path(temporal_images).mkdir(parents=True, exist_ok=True)
        fig_name = os.path.join(temporal_images, f"{column}.png")
        plt.savefig(fig_name, dpi=100)
        # plt.show()

    print("")
    print(f"Number of analyzed cells: {ca_cal.analyzed_cell_number(sheet)}.")

    for folder, subfolders, files in os.walk(converted_route):
        for subfolder in subfolders:
            if subfolder == non_space_filename:
                wb = openpyxl.load_workbook(file_route)
                sheet = wb.active
                for temporal_image in os.listdir(subfolder):
                    temporal_image_route = os.path.join(
                        converted_route, subfolder, temporal_image)
                    img = openpyxl.drawing.image.Image(
                        temporal_image_route)
                    img_cell = sheet.cell(
                        row=row_num_plots, column=column_num_plots).coordinate
                    img.anchor = img_cell
                    sheet.add_image(img)
                    row_num_plots += 17
                filename = f"{non_space_filename}_analyzed.xlsx"

                final_file_max_row = sheet.max_row
                final_file_max_column = sheet.max_column

                for row in range(1, final_file_max_row + 1):
                    for column in range(1, final_file_max_column):
                        if sheet.cell(row=row, column=column).value:
                            float_formating = sheet.cell(
                                row=row, column=column).coordinate
                            if type(sheet[float_formating].value) == float:
                                ca_ex_st.style_number(float_formating, sheet)
                            elif type(sheet[float_formating].value) == int:
                                ca_ex_st.style_time(float_formating, sheet)

                for row in range(1, final_file_max_row + 1):
                    if sheet.cell(row=row, column=column_num_data + 1).value:
                        parameter_sample = sheet.cell(
                            row=row, column=column_num_data + 1).coordinate
                        ca_ex_st.style_oscillation_calculated_parameters_header(
                            parameter_sample, sheet)

                ca_ex_st.colum_max_widths_oscillation_calcium_experiments(
                    sheet, column_num_data)
                wb.save(filename)
                os.remove(file_route)
                shutil.rmtree(subfolder)
                print(f"{filename} has been saved.")


def imaging_ca_analysis_PCA_multiple(adquisiton_time, keyword, peak_amplitude, peak_longitude, path_folder, time_initial_linregress, y_max_value, y_min_value):
    file_route = Path(path_folder)
    os.chdir(file_route)
    for folder, _, files in os.walk(file_route):
        for file in files:
            route = os.path.join(folder, file)
            imaging_ca_analysis_PCA(adquisiton_time, keyword, peak_amplitude,
                                    peak_longitude, route, time_initial_linregress, y_max_value, y_min_value)
